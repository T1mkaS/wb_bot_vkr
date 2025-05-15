import io
import json
import os
import time
import traceback
import uuid
from datetime import date, timedelta, datetime
import re
from aiogram import F
import pandas as pd
from aiogram import Router, types
from aiogram.enums import ParseMode
from aiogram.types import Message, CallbackQuery, FSInputFile, BufferedInputFile
from aiogram.filters import Command
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.utils.markdown import hbold
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from openpyxl.reader.excel import load_workbook
from aiogram.types import Message, Document
from openpyxl.styles import Alignment
from wb_assistance_bot.db.users import users_db
from wb_assistance_bot.tg_bot.bot import bot
from wb_assistance_bot.wb.exceptions import UnauthorizedException
from wb_assistance_bot.wb.financial_report.report_counter_no_api import report_counter
import wb_assistance_bot.tg_bot.keyboards.financial_report_kb as financial_report_kb
import wb_assistance_bot.tg_bot.prepared_text.financial_report_text as financial_report_text
from wb_assistance_bot.wb.financial_report.report_counter_with_api import Report
from wb_assistance_bot.wb.financial_report import report_counter_with_file
from wb_assistance_bot.wb.financial_report import storage_costs
from wb_assistance_bot.wb.financial_report import acceptance_cost
from wb_assistance_bot.wb.financial_report import advertising_expanses
from wb_assistance_bot.wb.financial_report import advertising_info
from wb_assistance_bot.wb.financial_report import products_info

router = Router()


def auto_adjust_columns(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(wrap_text=True)

        ws.column_dimensions[column].width = max_length + 2

    wb.save(file_path)


def auto_adjust_columns_in_memory(file_stream):
    wb = load_workbook(file_stream)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(
                    wrapText=True,
                    vertical='center',
                    horizontal='center',
                    textRotation=0,
                    indent=0,
                    shrinkToFit=False
                )

    for col in ws.columns:
        first_cell = col[0]
        column_letter = first_cell.column_letter
        current_width = ws.column_dimensions[column_letter].width

        if current_width is None:
            header_value = str(first_cell.value) if first_cell.value else ""
            current_width = len(header_value) + 2

        ws.column_dimensions[column_letter].width = min(current_width * 1.5, 50)

    file_stream.seek(0)
    wb.save(file_stream)
    file_stream.seek(0)


def get_last_full_weeks(n=4):
    today = date.today()
    current_monday = today - timedelta(days=today.weekday())
    result = []

    for i in range(n, 0, -1):
        start = current_monday - timedelta(weeks=i)
        end = start + timedelta(days=6)
        week_str = f"{start.strftime('%d.%m')}-{end.strftime('%d.%m')}"
        result.append(week_str)

    return result


def parse_range_to_rfc3339(week_range: str):
    start_str, end_str = week_range.split('-')

    today = date.today()

    start_date = datetime.strptime(start_str + f".{today.year}", "%d.%m.%Y").date()

    if start_date > today:
        start_date = datetime.strptime(start_str + f".{today.year - 1}", "%d.%m.%Y").date()

    end_date = datetime.strptime(end_str + f".{start_date.year}", "%d.%m.%Y").date()

    if end_date < start_date:
        end_date = datetime.strptime(end_str + f".{today.year}", "%d.%m.%Y").date()

    date_from = start_date.isoformat()
    date_to = end_date.isoformat()

    return date_from, date_to


class Form(StatesGroup):
    wait_for_key_finance = State()
    wait_for_report = State()
    wait_for_costs = State()


@router.message(Command("financial_report"))
async def supply_notifier(message: Message):
    await message.answer(financial_report_text.main_menu,
                         reply_markup=financial_report_kb.kb,
                         parse_mode=ParseMode.HTML)


@router.callback_query(lambda c: c.data == "change_finance_api_key")
async def changing_api_key(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.message.edit_text(financial_report_text.instruction_api_for_finance,
                                           parse_mode=ParseMode.HTML)
    await callback_query.message.edit_reply_markup(reply_markup=financial_report_kb.kb1)
    await state.update_data({'message_id': callback_query.message.message_id})
    await state.set_state(Form.wait_for_key_finance)


@router.message(Form.wait_for_key_finance)
async def change_key(message: Message, state: FSMContext):
    await users_db.set_finance_api(message.from_user.id,
                                   message.text)
    state_data = await state.get_data()
    await bot.delete_message(chat_id=message.chat.id, message_id=state_data['message_id'])
    await message.delete()
    await bot.send_message(chat_id=message.chat.id,
                           text="✅ Ключ успешно сохранен")
    await bot.send_message(chat_id=message.chat.id, text=financial_report_text.main_menu,
                           reply_markup=financial_report_kb.kb,
                           parse_mode=ParseMode.HTML)
    await state.clear()


@router.callback_query(lambda c: c.data == "close_financial_settings")
async def close_financial_settings(callback_query: CallbackQuery):
    await callback_query.message.delete()


@router.callback_query(lambda c: c.data == "back_to_finance_menu")
async def back_to_menu(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.message.edit_text(financial_report_text.main_menu,
                                           parse_mode=ParseMode.HTML)
    await callback_query.message.edit_reply_markup(reply_markup=financial_report_kb.kb)
    await state.clear()


@router.callback_query(lambda c: c.data == "about_api_finance")
async def close_financial_settings(callback_query: CallbackQuery):
    await callback_query.message.edit_text(text=financial_report_text.about_api_text, parse_mode=ParseMode.HTML)
    await callback_query.message.edit_reply_markup(reply_markup=financial_report_kb.kb2)


@router.callback_query(lambda c: c.data == "back_to_api_finance")
async def about_api(callback_query: CallbackQuery):
    await callback_query.message.edit_text(financial_report_text.instruction_api_for_finance, parse_mode=ParseMode.HTML)
    await callback_query.message.edit_reply_markup(reply_markup=financial_report_kb.kb1)


@router.callback_query(lambda c: c.data == "finance_no_api")
async def report_no_api(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.message.edit_text(financial_report_text.finance_no_api, parse_mode=ParseMode.HTML)
    await callback_query.message.edit_reply_markup(reply_markup=financial_report_kb.kb3)
    await state.update_data({'message_id': callback_query.message.message_id})
    await state.set_state(Form.wait_for_report)


@router.message(Form.wait_for_report)
async def waiting_for_report(message: Message, state: FSMContext):
    if not os.path.exists('reports'):
        os.makedirs('reports')
    try:
        document = message.document
        file_info = await bot.get_file(document.file_id)
        file_path = file_info.file_path
        unique_filename = f"{uuid.uuid4().hex}_{document.file_name}"
        filename = os.path.join('reports', unique_filename)
        await bot.download_file(file_path, filename)
        tax_rate = await users_db.get_tax_rate(message.from_user.id)
        tax_system = await users_db.get_tax_system(message.from_user.id)
        report = report_counter(filename, tax_rate, tax_system)
        await bot.send_document(chat_id=message.chat.id, document=types.FSInputFile(report, filename="report.xlsx"),
                                caption="Вот ваш отчет 📄")
        state_data = await state.get_data()
        await bot.delete_message(chat_id=message.chat.id, message_id=state_data['message_id'])
        await state.clear()
    except:
        await bot.send_message(chat_id=message.chat.id, text="Неправильный формат файла")


@router.callback_query(lambda c: c.data == "finance_with_api")
async def report_with_api(callback_query: CallbackQuery, state: FSMContext):
    fin_api = await users_db.get_finance_api(callback_query.from_user.id)
    print(1)
    if fin_api != "new_user":
        await callback_query.message.edit_text(financial_report_text.finance_with_api, parse_mode=ParseMode.HTML)
        weeks = get_last_full_weeks()
        kb = InlineKeyboardBuilder()
        for week in weeks:
            kb.button(text=week, callback_data=week)
        kb.adjust(1)
        kb = kb.as_markup()
        await callback_query.message.edit_reply_markup(reply_markup=kb)
    else:
        await callback_query.answer("Вы еще не ввели API-ключ для работы с разделом")


@router.callback_query(lambda c: c.data == "products_cost")
async def products_cost(callback_query: CallbackQuery, state: FSMContext):
    fin_api = await users_db.get_finance_api(callback_query.from_user.id)
    if fin_api != "new_user":
        await callback_query.message.edit_text(financial_report_text.products_costs_info, parse_mode=ParseMode.HTML)
        file_path = f'prod_costs/{callback_query.from_user.id}.xlsx'
        df_costs = pd.read_excel(file_path)
        if not df_costs.empty:
            try:
                products = products_info.Products(key=fin_api).get_products()
                df_all_goods = pd.DataFrame(columns=["Артикул поставщика", "Артикул WB", "Себестоимость"])
                for product in products['cards']:
                    prod_nmID = product['nmID']
                    prod_article = product['vendorCode']
                    new_product = {
                        "Артикул поставщика": prod_article,
                        "Артикул WB": prod_nmID,
                        "Себестоимость": 0
                    }
                    df_all_goods = pd.concat([df_all_goods, pd.DataFrame([new_product])], ignore_index=True)

                new_rows = df_all_goods[~df_all_goods["Артикул поставщика"].isin(df_costs["Артикул поставщика"])]

                df_costs = pd.concat([df_costs, new_rows], ignore_index=True)

                df_costs.to_excel(file_path, index=False)
                auto_adjust_columns(file_path)

            except UnauthorizedException:
                await callback_query.answer("API-ключ некорректный")
                await callback_query.message.delete()
            except Exception as e:
                print(traceback.format_exc())
                if str(e) == "Много запросов":
                    await callback_query.answer("WB не успевают предоставлять данные, подождите")
                elif str(e) != "Wb лажает":
                    await callback_query.answer("API-ключ некорректный")
                    await callback_query.message.delete()
        else:
            try:
                products = products_info.Products(key=fin_api).get_products()
                for product in products['cards']:
                    prod_nmID = product['nmID']
                    prod_article = product['vendorCode']
                    new_product = {
                        "Артикул поставщика": prod_article,
                        "Артикул WB": prod_nmID,
                        "Себестоимость": 0
                    }
                    df_costs = pd.concat([df_costs, pd.DataFrame([new_product])], ignore_index=True)
                df_costs.to_excel(file_path, index=False)
                auto_adjust_columns(file_path)

            except UnauthorizedException:
                await callback_query.answer("API-ключ некорректный")
                await callback_query.message.delete()
            except Exception as e:
                print(traceback.format_exc())
                if str(e) == "Много запросов":
                    await callback_query.answer("WB не успевают предоставлять данные, подождите")
                elif str(e) != "Wb лажает":
                    await callback_query.answer("API-ключ некорректный")
                    await callback_query.message.delete()

        document = FSInputFile(file_path)

        msg = await callback_query.message.answer_document(document)
        sent_message_id = msg.message_id

        await state.update_data({'message_id': callback_query.message.message_id, 'sent_message_id': sent_message_id})
        await state.set_state(Form.wait_for_costs)
    else:
        await callback_query.answer("Вы еще не ввели API-ключ для работы с разделом")


@router.message(Form.wait_for_costs, F.document)
async def new_costs(message: Message, state: FSMContext):
    state_data = await state.get_data()
    document: Document = message.document
    if not document.file_name.endswith('.xlsx'):
        await message.answer("Пожалуйста, отправьте Excel-файл с расширением .xlsx 📄")
    else:
        user_id = message.from_user.id
        file_path = f"prod_costs/{user_id}.xlsx"
        file = await bot.get_file(document.file_id)
        await bot.download_file(file.file_path, destination=file_path)

        await message.delete()
        await bot.delete_message(chat_id=message.chat.id, message_id=state_data['message_id'])
        await bot.delete_message(chat_id=message.chat.id, message_id=state_data['sent_message_id'])
        await bot.send_message(chat_id=message.chat.id,
                               text="✅Себестоимости обновлены")
        await bot.send_message(chat_id=message.chat.id, text=financial_report_text.main_menu,
                               reply_markup=financial_report_kb.kb,
                               parse_mode=ParseMode.HTML)
        await state.clear()


@router.callback_query()
async def catch_all_callbacks(callback: types.CallbackQuery):
    data = callback.data
    if re.match(r"^\d{2}\.\d{2}-\d{2}\.\d{2}$", data):
        date_from, date_to = parse_range_to_rfc3339(data)
        key = await users_db.get_finance_api(callback.from_user.id)
        rep_counter = Report(key=key)
        try:
            report = rep_counter.get_report(date_from, date_to)
            tax_rate = await users_db.get_tax_rate(callback.from_user.id)
            tax_system = await users_db.get_tax_system(callback.from_user.id)
            await callback.message.edit_text(f"Идет рассчет отчета {data}...")
            df = pd.DataFrame(report)
            df.rename(columns={
                "doc_type_name": "Тип документа",
                "supplier_oper_name": "Обоснование для оплаты",
                "sa_name": "Артикул поставщика",
                "quantity": "Кол-во",
                "retail_amount": "Вайлдберриз реализовал Товар (Пр)",
                "ppvz_for_pay": "К перечислению Продавцу за реализованный Товар",
                "delivery_rub": "Услуги по доставке товара покупателю",
                "penalty": "Общая сумма штрафов",
                "nm_id": "Артикул WB"
            }, inplace=True)

            counted_report = report_counter_with_file.report_counter(df, tax_rate, tax_system)
            storager = storage_costs.Storage(key)
            acceptancer = acceptance_cost.Acceptance(key)

            storage_expanses = storager.get_storage_cost(date_from, date_to)
            df_storage = pd.DataFrame(storage_expanses)
            df_storage = df_storage.groupby("vendorCode", as_index=False)["warehousePrice"].sum()
            df_storage.rename(columns={
                "vendorCode": "Артикул поставщика",
                "warehousePrice": "Сумма хранения"
            }, inplace=True)
            df_storage['Артикул поставщика'] = df_storage['Артикул поставщика'].str.lower()
            report = counted_report.merge(df_storage, on='Артикул поставщика', how='left')

            time.sleep(60)

            acceptance_expanses = acceptancer.get_acceptance_cost(date_from, date_to)
            df_acceptance = pd.DataFrame(acceptance_expanses)
            if not df_acceptance.empty:
                df_acceptance = df_acceptance.groupby("nmID", as_index=False)["total"].sum()
                df_acceptance.rename(columns={
                    "nmID": "Артикул WB",
                    "total": "Сумма платной приемки"
                }, inplace=True)
                report = report.merge(df_acceptance, on='Артикул WB', how='left')
            else:
                report["Сумма платной приемки"] = 0
            cols = list(report.columns)
            cols.remove("Артикул WB")
            cols.insert(1, "Артикул WB")
            report = report[cols]
            # report.to_excel("отчет.xlsx", index=False)

            advertising_campaigns = advertising_expanses.AdvertieseExpanses(key)
            campaigns = advertising_campaigns.get_advertising_cost(date_from, date_to)
            ids = []
            for campaign in campaigns:
                campaign_id = int(campaign["advertId"])
                ids.append(campaign_id)
            ids = list(set(ids))
            if len(ids) != 0:
                advertising_information = advertising_info.AdvertieseInfo(key)
                campaigns_with_articles = advertising_information.get_advertising_information(ids)
                total = {}
                for campaign in campaigns:
                    campaign_id = int(campaign["advertId"])
                    if campaign_id in total:
                        total[campaign_id] += int(campaign["updSum"])
                    else:
                        total[campaign_id] = int(campaign["updSum"])
                id_article = {}
                for campaign in campaigns_with_articles:
                    advert_id = campaign["advertId"]
                    if "unitedParams" in campaign.keys():
                        nmId = campaign["unitedParams"][0]["nms"]
                    else:
                        try:
                            nmId = campaign["autoParams"][0]["nms"]
                        except:
                            nmId = campaign["autoParams"]["nms"]
                    id_article[advert_id] = nmId[0]

                article_sum = {}
                for campaign_id, article in id_article.items():
                    amount = total.get(campaign_id, 0)
                    article_sum[article] = article_sum.get(article, 0) + amount
                df_advertising_expanses = pd.DataFrame([
                    {"Артикул WB": article, "Затраты на рекламу": total_sum}
                    for article, total_sum in article_sum.items()
                ])
                report = report.merge(df_advertising_expanses, on='Артикул WB', how='left')

            # работа с себестоимостью
            file_path = f'prod_costs/{callback.from_user.id}.xlsx'
            df_prod_costs = pd.read_excel(file_path)
            df_prod_costs['Артикул поставщика'] = df_prod_costs['Артикул поставщика'].astype(str).str.lower()

            merged_report = report.merge(df_prod_costs[['Артикул поставщика', 'Себестоимость']],
                                         on='Артикул поставщика',
                                         how='left')
            merged_report['Себестоимость'] = merged_report['Себестоимость'].fillna(0)
            merged_report['Себестоимость проданных товаров'] = merged_report['Себестоимость'] * merged_report[
                'Кол-во проданных - возвращенных товаров']
            merged_report.drop(columns=['Себестоимость'], inplace=True)
            columns = merged_report.columns.tolist()
            columns.remove('Себестоимость проданных товаров')
            columns.insert(5, 'Себестоимость проданных товаров')
            merged_report = merged_report[columns]

            merged_report.fillna(0, inplace=True)

            if tax_system == "incomes_expenses":
                merged_report["Налог"] = (merged_report["Сумма продажи WB (продажи - возвраты)"] -
                                          merged_report["Сумма логистики"] -
                                          merged_report["Сумма комиссии"] -
                                          merged_report["Общая сумма штрафов"] -
                                          merged_report["Сумма хранения"] -
                                          merged_report["Сумма платной приемки"] -
                                          merged_report["Затраты на рекламу"]) * tax_rate / 100

            merged_report["Чистая прибыль"] = (merged_report["Сумма продажи WB (продажи - возвраты)"] -
                                               merged_report["Сумма логистики"] -
                                               merged_report["Сумма комиссии"] -
                                               merged_report["Общая сумма штрафов"] -
                                               merged_report["Сумма хранения"] -
                                               merged_report["Сумма платной приемки"] -
                                               merged_report["Затраты на рекламу"] -
                                               merged_report["Себестоимость проданных товаров"] -
                                               merged_report["Налог"]
                                               )
            merged_report.drop(columns=['Cумма к перечислению продавцу с учетом всех расходов вб'], inplace=True)

            file_stream = io.BytesIO()
            merged_report.to_excel(file_stream, index=False)
            auto_adjust_columns_in_memory(file_stream)
            file_stream.seek(0)
            input_file = BufferedInputFile(file_stream.read(), filename="Отчет.xlsx")
            await bot.send_document(
                chat_id=callback.message.chat.id,
                document=input_file,
                caption=f"Рассчитанный отчет {data}"
            )
            await callback.message.delete()

        except UnauthorizedException:
            await callback.answer("API-ключ некорректный")
            await callback.message.delete()
        except Exception as e:
            print(traceback.format_exc())
            if str(e) == "Много запросов":
                await callback.answer("WB не успевают предоставлять данные, подождите")
            elif str(e) != "Wb лажает":
                await callback.answer("API-ключ некорректный")
                await callback.message.delete()
